from flask import Flask, render_template, request, send_file, jsonify, Response
import io
import os
import pdf2image
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
import re
import json

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = '/tmp/suez_uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Excel formatting constants
HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column widths as per specification
COLUMN_WIDTHS = {
    'A': 8,   # Week
    'B': 18,  # Type
    'C': 12,  # Date
    'D': 12,  # Day
    'E': 10,  # IN
    'F': 10,  # OUT
    'G': 10,  # Nett
    'H': 14,  # Street/Litter
    'I': 14,  # Flytip Monthly
    'J': 10,  # Compost
    'K': 18   # Ticket No
}


def enhance_image(image):
    """Enhance image for better OCR"""
    # Convert to grayscale
    image = image.convert('L')

    # Increase contrast
    enhancer = ImageEnhance.Contrast(image)
    image = enhancer.enhance(2.0)

    # Sharpen
    image = image.filter(ImageFilter.SHARPEN)

    # Threshold to black and white
    threshold = 128
    image = image.point(lambda p: 255 if p > threshold else 0)

    return image


def determine_waste_type(text):
    """Determine waste type from CONTRACT DESC field"""
    upper_text = text.upper()

    # Look for CONTRACT DESC field specifically
    contract_match = re.search(r'CONTRACT\s*DESC[:\s]*([A-Z0-9\s\-]+)', upper_text)
    contract_desc = contract_match.group(1) if contract_match else upper_text

    # Apply mapping logic as per specification
    if 'SWEEPINGS' in contract_desc or 'SWEEPING' in contract_desc:
        return 'Street Sweepings'
    elif 'LITTER' in contract_desc:
        return 'Street Litter'
    elif any(word in contract_desc for word in ['BULKY', 'FLY', 'MUNICIPAL', 'MIXED',
                                                  'BIODEGRADABLE', 'KITCHEN', 'TYRES',
                                                  'FRIDGES', 'FREEZERS']):
        return 'Bulky Waste'
    else:
        return 'Unknown'


def extract_tickets_from_text(text, page_number):
    """Extract ticket data from OCR text - supports multiple tickets per page"""
    tickets = []
    issues = []

    # Clean up text
    text = text.replace('|', 'I').replace('`', '').replace("'", '')

    # Find all ticket numbers on the page
    ticket_pattern = r'(\d{4}[\s\-]?[PpBbRr][\s\-]?\d{9})'
    ticket_matches = list(re.finditer(ticket_pattern, text))

    if not ticket_matches:
        # Try alternative pattern with Ticket No: label
        ticket_matches = list(re.finditer(r'Ticket\s*No[:\s]*(\d{4}[\s\-]?[PpBbRr][\s\-]?\d{9})', text, re.IGNORECASE))

    if not ticket_matches:
        issues.append(f"Page {page_number}: No ticket number found")
        return tickets, issues

    # For each ticket found, extract associated data
    for i, ticket_match in enumerate(ticket_matches):
        ticket_data = extract_single_ticket(text, ticket_match, page_number, i + 1)
        if ticket_data['ticket']:
            tickets.append(ticket_data['ticket'])
        if ticket_data['issues']:
            issues.extend(ticket_data['issues'])

    return tickets, issues


def extract_single_ticket(text, ticket_match, page_number, ticket_index):
    """Extract a single ticket's data from text"""
    ticket = None
    issues = []

    # Extract ticket number (last 9 digits)
    ticket_full = ticket_match.group(1).replace(' ', '').replace('-', '').upper()
    ticket_full = re.sub(r'[BR]', 'P', ticket_full)
    ticket_no = ticket_full[-9:]

    # Validate ticket number format (9 digits)
    if not re.match(r'^\d{9}$', ticket_no):
        issues.append(f"Page {page_number}: Invalid ticket number format '{ticket_no}'")
        return {'ticket': None, 'issues': issues}

    # Find date - flexible year pattern
    date_match = re.search(r'Date[:\s]*(\d{1,2})[\s\-]([A-Za-z]{3})[\s\-](\d{4})', text, re.IGNORECASE)
    if not date_match:
        # Fallback: search for date pattern anywhere
        date_match = re.search(r'(\d{1,2})[\s\-]([A-Za-z]{3})[\s\-](\d{4})', text)

    if not date_match:
        issues.append(f"Page {page_number}: No date found for ticket {ticket_no}")
        return {'ticket': None, 'issues': issues}

    day = date_match.group(1).zfill(2)
    month = date_match.group(2)
    year = date_match.group(3)

    try:
        date_obj = datetime.strptime(f"{day}-{month}-{year}", "%d-%b-%Y")
    except ValueError:
        issues.append(f"Page {page_number}: Invalid date format for ticket {ticket_no}")
        return {'ticket': None, 'issues': issues}

    # Find weights - look for GROSS, TARE, NET patterns
    gross_match = re.search(r'GROSS[^\d]{0,20}(\d{3,5})', text, re.IGNORECASE)
    tare_match = re.search(r'TARE[^\d]{0,20}(\d{3,5})', text, re.IGNORECASE)
    net_match = re.search(r'NET[^\d]{0,20}(\d{2,5})', text, re.IGNORECASE)

    gross = int(gross_match.group(1)) if gross_match else None
    tare = int(tare_match.group(1)) if tare_match else None
    net = int(net_match.group(1)) if net_match else None

    # Validate and recalculate net if needed
    if gross is None or gross == 0:
        issues.append(f"Page {page_number}: Missing gross weight for ticket {ticket_no}")
        return {'ticket': None, 'issues': issues}

    if tare is None or tare == 0:
        issues.append(f"Page {page_number}: Missing tare weight for ticket {ticket_no}")
        return {'ticket': None, 'issues': issues}

    # Calculate expected net
    calculated_net = gross - tare

    # If net is missing or doesn't match calculation, use calculated value
    if net is None or abs(net - calculated_net) > 1:
        if net is not None:
            issues.append(f"Page {page_number}: Net weight mismatch for ticket {ticket_no}: "
                         f"found {net}, calculated {calculated_net}. Using calculated value.")
        net = calculated_net

    # Final validation
    if net <= 0:
        issues.append(f"Page {page_number}: Invalid net weight ({net}) for ticket {ticket_no}")
        return {'ticket': None, 'issues': issues}

    if tare > gross:
        issues.append(f"Page {page_number}: Tare ({tare}) exceeds gross ({gross}) for ticket {ticket_no}")
        return {'ticket': None, 'issues': issues}

    # Determine waste type
    waste_type = determine_waste_type(text)
    if waste_type == 'Unknown':
        issues.append(f"Page {page_number}: Unknown waste type for ticket {ticket_no}")

    ticket = {
        'ticket_no': ticket_no,
        'date': date_obj,
        'day': date_obj.strftime('%A'),
        'gross': gross,
        'tare': tare,
        'net': net,
        'type': waste_type,
        'page': page_number
    }

    return {'ticket': ticket, 'issues': issues}


def create_excel_workbook(tickets):
    """Create formatted Excel workbook"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suez Glossop"  # Sheet name as per specification

    # Set column widths
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Headers
    headers = ["Week", "Type", "Date", "Day", "IN", "OUT", "Nett",
               "Street/Litter", "Flytip Monthly", "Compost", "Ticket No"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # Data rows
    for row_num, ticket in enumerate(tickets, 2):
        # Week (blank)
        cell = ws.cell(row_num, 1)
        cell.value = None
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # Type
        cell = ws.cell(row_num, 2)
        cell.value = ticket['type']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # Date
        cell = ws.cell(row_num, 3)
        cell.value = ticket['date']
        cell.number_format = 'DD/MM/YYYY'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # Day
        cell = ws.cell(row_num, 4)
        cell.value = ticket['day']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # IN (Gross weight)
        cell = ws.cell(row_num, 5)
        cell.value = ticket['gross']
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # OUT (Tare weight)
        cell = ws.cell(row_num, 6)
        cell.value = ticket['tare']
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # Nett (Net weight)
        cell = ws.cell(row_num, 7)
        cell.value = ticket['net']
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # Street/Litter (blank)
        cell = ws.cell(row_num, 8)
        cell.value = None
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # Flytip Monthly (blank)
        cell = ws.cell(row_num, 9)
        cell.value = None
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # Compost (blank)
        cell = ws.cell(row_num, 10)
        cell.value = None
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # Ticket No (right-aligned)
        cell = ws.cell(row_num, 11)
        cell.value = ticket['ticket_no']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = THIN_BORDER

    return wb


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
def process_pdf():
    if 'pdf' not in request.files:
        return jsonify({'error': 'No PDF file uploaded'}), 400

    pdf_file = request.files['pdf']
    if pdf_file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    # Read file bytes BEFORE entering generator (file closes after request context)
    pdf_bytes = pdf_file.read()

    def generate():
        try:
            # Send initial status
            yield f"data: {json.dumps({'status': 'converting', 'message': 'Converting PDF to images...'})}\n\n"

            # Convert PDF to images
            try:
                images = pdf2image.convert_from_bytes(pdf_bytes, dpi=200)
            except Exception as e:
                yield f"data: {json.dumps({'status': 'error', 'message': f'PDF conversion failed: {str(e)}'})}\n\n"
                return

            all_tickets = []
            all_issues = []
            total_pages = len(images)

            if total_pages == 0:
                yield f"data: {json.dumps({'status': 'error', 'message': 'No pages found in PDF'})}\n\n"
                return

            for page_num, image in enumerate(images, 1):
                # Enhance image
                enhanced = enhance_image(image)

                # Perform OCR
                text = pytesseract.image_to_string(enhanced)

                # Extract tickets (supports multiple per page)
                tickets, issues = extract_tickets_from_text(text, page_num)
                all_tickets.extend(tickets)
                all_issues.extend(issues)

                yield f"data: {json.dumps({'status': 'processing', 'page': page_num, 'total': total_pages, 'found': len(all_tickets)})}\n\n"

            if not all_tickets:
                error_msg = 'No tickets found'
                if all_issues:
                    error_msg += '. Issues: ' + '; '.join(all_issues[:5])
                    if len(all_issues) > 5:
                        error_msg += f' ... and {len(all_issues) - 5} more'
                yield f"data: {json.dumps({'status': 'error', 'message': error_msg})}\n\n"
                return

            # Sort by date
            all_tickets.sort(key=lambda x: x['date'])

            # Create Excel workbook with full formatting
            wb = create_excel_workbook(all_tickets)

            # Save file with specified filename
            filename = "Suez_Glossop_Tickets.xlsx"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            wb.save(filepath)

            # Prepare response with issues report
            response_data = {
                'status': 'complete',
                'tickets': len(all_tickets),
                'filename': filename
            }

            if all_issues:
                response_data['issues'] = all_issues
                response_data['pages_needing_review'] = list(set(
                    int(issue.split(':')[0].replace('Page ', ''))
                    for issue in all_issues if issue.startswith('Page ')
                ))

            yield f"data: {json.dumps(response_data)}\n\n"

        except Exception as e:
            yield f"data: {json.dumps({'status': 'error', 'message': str(e)})}\n\n"

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no'  # Disable nginx buffering
        }
    )


@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name=filename)
    return jsonify({'error': 'File not found'}), 404


if __name__ == '__main__':
    app.run(debug=True)
